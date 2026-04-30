"""
merge_all_countries.py
Merges water law judicial decisions from Brazil, Canada, and the Netherlands
into a single national + international CSV and XLSX dataset.

Unified schema:
  country | tribunal | court_name | case_id | title | date | year |
  case_type | chamber | judge | legal_area | summary | url
"""
import json, csv, sys, os
from pathlib import Path
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding='utf-8')

DL  = Path(os.environ.get('DATA_DIR', r'C:\Users\junio\Downloads'))
OUT_CSV  = DL / 'water_law_global.csv'
OUT_XLSX = DL / 'water_law_global.xlsx'

# ── Country colours for XLSX ──────────────────────────────────────────────────
COLORS = {
    'Brazil':      {'bg': 'DDEEFF', 'header': '1A5276'},  # blue
    'Canada':      {'bg': 'FEF9E7', 'header': '7D6608'},  # amber
    'Netherlands': {'bg': 'EAFAF1', 'header': '1E8449'},  # green
    'Brazil-Historical': {'bg': 'F5EEF8', 'header': '6C3483'},  # purple (TJSP 1997-2015)
}

# ── Unified fieldnames ────────────────────────────────────────────────────────
FIELDS = [
    'country', 'tribunal', 'court_name', 'case_id',
    'title', 'date', 'year', 'case_type', 'chamber',
    'judge', 'legal_area', 'summary', 'url',
]

# ── Normalizers per country ───────────────────────────────────────────────────

def normalize_brazil(c):
    ementa = (c.get('ementa', '') or '')
    return {
        'country':    'Brazil',
        'tribunal':   c.get('tribunal', ''),
        'court_name': f"Tribunal de Justiça do {c.get('estado', '')}",
        'case_id':    c.get('num_processo', '') or c.get('numero_processo', ''),
        'title':      '',
        'date':       c.get('data_julgamento', '') or c.get('data_publicacao', ''),
        'year':       str(c.get('ano', '')),
        'case_type':  c.get('classe', ''),
        'chamber':    c.get('camara_orgao', '') or c.get('orgao_julgador', ''),
        'judge':      c.get('relator', ''),
        'legal_area': 'Direito à Água / Saneamento',
        'summary':    ementa[:3000],
        'url':        c.get('url', ''),
    }

def normalize_canada(c):
    # Handles both CanLII scraper output and LDH (CA/A2AJ) output
    source = c.get('source', '')
    court_name = (
        c.get('court_name') or                        # CanLII scraper
        ('Access to Administrative Justice (A2AJ)'    # LDH fallback
         if 'A2AJ' in c.get('tribunal', '') else
         c.get('tribunal', ''))
    )
    return {
        'country':    'Canada',
        'tribunal':   c.get('tribunal', ''),
        'court_name': court_name,
        'case_id':    (c.get('citation') or c.get('source_id') or
                       c.get('docket', '')),
        'title':      c.get('title', ''),
        'date':       c.get('date', '') or c.get('decisionDate', ''),
        'year':       str(c.get('year', '')),
        'case_type':  '',
        'chamber':    '',
        'judge':      '',
        'legal_area': 'Water Law',
        'summary':    (c.get('snippet') or c.get('keywords') or '')[:3000],
        'url':        c.get('url', ''),
    }

def normalize_netherlands(c):
    return {
        'country':    'Netherlands',
        'tribunal':   c.get('court', ''),
        'court_name': c.get('court_name', ''),
        'case_id':    c.get('ecli', '') or c.get('zaak_nummer', ''),
        'title':      c.get('ecli', ''),
        'date':       c.get('decision_date', ''),
        'year':       str(c.get('year', '')),
        'case_type':  c.get('procedure', ''),
        'chamber':    '',
        'judge':      '',
        'legal_area': c.get('rechtsgebied', ''),
        'summary':    (c.get('inhoudsindicatie', '') or '')[:3000],
        'url':        c.get('url', ''),
    }

# ── Load each country ─────────────────────────────────────────────────────────
all_cases = []

def load_brazil():
    brazil = []
    court_files = [
        ('tjsp_cases_all.json',          None),
        ('tjsc_cases_2016_2026.json',    None),
        ('tjrr_cases_2016_2026.json',    None),
        ('tjac_cases_2016_2026.json',    None),
        ('tjpi_cases_2016_2026.json',    None),
        ('tjto_cases_2016_2026.json',    None),
        ('tjdft_cases_2016_2026.json',   None),
        ('tjrj_cases_2016_2026.json',    None),
    ]
    for fname, _ in court_files:
        fpath = DL / fname
        if not fpath.exists():
            print(f'  MISSING: {fname}')
            continue
        with open(fpath, encoding='utf-8') as f:
            data = json.load(f)
        court_cases = [normalize_brazil(c) for c in data]
        print(f'  {fname}: {len(court_cases)} cases')
        brazil.extend(court_cases)
    return brazil

def load_canada():
    fpath = DL / 'canada_water_law_2016_2026.json'
    if not fpath.exists():
        print('  MISSING: canada_water_law_2016_2026.json (run ldh_canada_scraper.py first)')
        return []
    with open(fpath, encoding='utf-8') as f:
        data = json.load(f)
    cases = [normalize_canada(c) for c in data]
    print(f'  canada_water_law_2016_2026.json: {len(cases)} cases')
    return cases

def normalize_tjsp_transformation(c):
    """Normalize TJSP Transformation Analysis 2021 record (1997-2015 historical cases)."""
    return {
        'country':    'Brazil',
        'tribunal':   'TJSP',
        'court_name': 'Tribunal de Justiça de São Paulo',
        'case_id':    str(c.get('case_id', '')),
        'title':      c.get('title', ''),
        'date':       c.get('date', ''),
        'year':       str(c.get('year', '')),
        'case_type':  c.get('case_type', ''),
        'chamber':    c.get('chamber', ''),
        'judge':      c.get('judge', ''),
        'legal_area': c.get('legal_area', ''),
        'summary':    c.get('summary', '')[:3000],
        'url':        c.get('url', ''),
        # pass-through qualitative fields for the historical sheet
        '_municipality':         c.get('municipality', ''),
        '_plaintiff':            c.get('plaintiff', ''),
        '_defendant':            c.get('defendant', ''),
        '_key_topic':            c.get('key_topic', ''),
        '_key_legislation':      c.get('key_legislation', ''),
        '_hr_language':          c.get('hr_language'),
        '_sust_language':        c.get('sust_language'),
        '_collective_claim':     c.get('collective_claim'),
        '_mp_involved':          c.get('mp_involved'),
        '_win':                  c.get('win'),
        '_mixed_result':         c.get('mixed_result'),
        '_governance_category':  c.get('governance_category', ''),
        '_language_category':    c.get('language_category', ''),
        '_notes':                c.get('notes', ''),
    }

def load_tjsp_transformation():
    fpath = DL / 'tjsp_transformation_analysis.json'
    if not fpath.exists():
        print('  (not available: tjsp_transformation_analysis.json)')
        return []
    with open(fpath, encoding='utf-8') as f:
        data = json.load(f)
    cases = [normalize_tjsp_transformation(c) for c in data]
    print(f'  tjsp_transformation_analysis.json: {len(cases)} cases (TJSP 1997-2015)')
    return cases

def load_netherlands():
    nl_files = [
        'netherlands_water_law_2016_2026.json',     # RvS + CBb + GHARL + HR (overnight scraper)
        'netherlands_district_courts_2016_2026.json', # 11 Rechtbanken + Gerechtshoven (district scraper)
        'netherlands_ldh_2016_2026.json',            # LDH: NL/Rechtspraak + SupremeCourt + CRvB
    ]
    all_nl = []
    seen_eclis = set()
    for fname in nl_files:
        fpath = DL / fname
        if not fpath.exists():
            print(f'  (not yet available: {fname})')
            continue
        with open(fpath, encoding='utf-8') as f:
            data = json.load(f)
        new = 0
        for c in data:
            ecli = c.get('ecli') or c.get('source_id') or c.get('title', '')[:60]
            if ecli in seen_eclis:
                continue
            # Year guard: Rechtspraak API date filter isn't 100% tight
            yr = c.get('year')
            if yr and not (2016 <= int(yr) <= 2026):
                continue
            seen_eclis.add(ecli)
            all_nl.append(normalize_netherlands(c))
            new += 1
        print(f'  {fname}: {new} cases (deduped)')
    return all_nl

print('Loading Brazil...')
brazil_cases = load_brazil()

print('Loading TJSP Transformation Analysis (historical 1997-2015)...')
tjsp_hist_cases = load_tjsp_transformation()

print('Loading Canada...')
canada_cases = load_canada()

print('Loading Netherlands...')
nl_cases = load_netherlands()

all_cases = brazil_cases + tjsp_hist_cases + canada_cases + nl_cases
# Sort by country then date descending
all_cases.sort(key=lambda x: (x['country'], x['date'] or ''), reverse=False)

print(f'\nTotal: {len(all_cases)} cases across {len(set(c["country"] for c in all_cases))} countries')

# ── Save CSV ──────────────────────────────────────────────────────────────────
with open(OUT_CSV, 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=FIELDS)
    writer.writeheader()
    for row in all_cases:
        writer.writerow({k: row.get(k, '') for k in FIELDS})
print(f'CSV saved: {OUT_CSV}')

# ── Save XLSX ─────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def make_header_style(hex_color):
        return {
            'font':  Font(bold=True, color='FFFFFF', size=10),
            'fill':  PatternFill('solid', fgColor=hex_color),
            'align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        }

    def style_header_row(ws, col_count, hex_color):
        style = make_header_style(hex_color)
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font  = style['font']
            cell.fill  = style['fill']
            cell.alignment = style['align']
        ws.row_dimensions[1].height = 28

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def freeze_and_filter(ws):
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    COL_WIDTHS = [13, 10, 30, 28, 35, 13, 6, 20, 25, 22, 20, 60, 40]
    HUMAN_HEADERS = [
        'Country', 'Tribunal', 'Court Name', 'Case ID',
        'Title', 'Date', 'Year', 'Case Type', 'Chamber',
        'Judge / Relator', 'Legal Area', 'Summary', 'URL',
    ]

    # ── Sheet 1: All Cases ────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = 'All Cases'
    ws_all.append(HUMAN_HEADERS)
    style_header_row(ws_all, len(FIELDS), '2C3E50')
    set_col_widths(ws_all, COL_WIDTHS)

    country_fills = {c: PatternFill('solid', fgColor=COLORS[c]['bg']) for c in COLORS}

    for row in all_cases:
        vals = [row.get(f, '') for f in FIELDS]
        ws_all.append(vals)
        country = row.get('country', '')
        if country in country_fills:
            fill = country_fills[country]
            for col in range(1, len(FIELDS) + 1):
                ws_all.cell(row=ws_all.max_row, column=col).fill = fill

    freeze_and_filter(ws_all)

    # ── Per-country sheets ────────────────────────────────────────────────────
    country_groups = {
        'Brazil':      brazil_cases,
        'Canada':      canada_cases,
        'Netherlands': nl_cases,
    }

    for country, cases in country_groups.items():
        if not cases:
            continue
        ws = wb.create_sheet(country)
        ws.append(HUMAN_HEADERS)
        style_header_row(ws, len(FIELDS), COLORS[country]['header'])
        set_col_widths(ws, COL_WIDTHS)
        fill = country_fills[country]
        for row in cases:
            vals = [row.get(f, '') for f in FIELDS]
            ws.append(vals)
            for col in range(1, len(FIELDS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
        freeze_and_filter(ws)

    # ── TJSP Historical sheet (1997-2015) with rich qualitative fields ────────
    if tjsp_hist_cases:
        HIST_FIELDS = FIELDS + [
            '_municipality', '_plaintiff', '_defendant', '_key_topic',
            '_key_legislation', '_hr_language', '_sust_language',
            '_collective_claim', '_mp_involved', '_win', '_mixed_result',
            '_governance_category', '_language_category', '_notes',
        ]
        HIST_HEADERS = HUMAN_HEADERS + [
            'Municipality', 'Plaintiff', 'Defendant', 'Key Topic',
            'Key Legislation', 'HR Language', 'Sust. Language',
            'Collective Claim', 'MP Involved', 'Win', 'Mixed Result',
            'Governance Category', 'Language Category', 'Notes',
        ]
        ws_h = wb.create_sheet('TJSP Historical (1997-2015)')
        ws_h.append(HIST_HEADERS)
        style_header_row(ws_h, len(HIST_FIELDS), COLORS['Brazil-Historical']['header'])
        hist_widths = COL_WIDTHS + [18, 22, 22, 40, 30, 10, 10, 14, 10, 6, 10, 28, 22, 40]
        set_col_widths(ws_h, hist_widths)
        hist_fill = PatternFill('solid', fgColor=COLORS['Brazil-Historical']['bg'])
        for row in tjsp_hist_cases:
            vals = [row.get(f, '') for f in HIST_FIELDS]
            ws_h.append(vals)
            for col in range(1, len(HIST_FIELDS) + 1):
                ws_h.cell(row=ws_h.max_row, column=col).fill = hist_fill
        freeze_and_filter(ws_h)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    ws_sum.column_dimensions['A'].width = 16
    ws_sum.column_dimensions['B'].width = 14
    ws_sum.column_dimensions['C'].width = 14

    # Header
    ws_sum['A1'] = 'Water Law Dataset — Global Summary'
    ws_sum['A1'].font = Font(bold=True, size=13, color='2C3E50')
    ws_sum.merge_cells('A1:C1')
    ws_sum.row_dimensions[1].height = 24

    ws_sum.append([])
    ws_sum.append(['Country', 'Cases', '% of Total'])
    for col in range(1, 4):
        c = ws_sum.cell(row=3, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2C3E50')
        c.alignment = Alignment(horizontal='center')

    total = len(all_cases)
    summary_groups = {
        'Brazil (2016-2026)':      brazil_cases,
        'Brazil Historical TJSP':  tjsp_hist_cases,
        'Canada':                  canada_cases,
        'Netherlands':             nl_cases,
    }
    fill_map = {
        'Brazil (2016-2026)':      country_fills['Brazil'],
        'Brazil Historical TJSP':  PatternFill('solid', fgColor=COLORS['Brazil-Historical']['bg']),
        'Canada':                  country_fills['Canada'],
        'Netherlands':             country_fills['Netherlands'],
    }
    for label, cases in summary_groups.items():
        n = len(cases)
        pct = f'{n/total*100:.1f}%' if total else '0%'
        ws_sum.append([label, n, pct])
        row_idx = ws_sum.max_row
        for col in range(1, 4):
            ws_sum.cell(row=row_idx, column=col).fill = fill_map[label]

    ws_sum.append([])
    ws_sum.append(['TOTAL', total, '100%'])
    total_row = ws_sum.max_row
    for col in range(1, 4):
        c = ws_sum.cell(row=total_row, column=col)
        c.font = Font(bold=True)

    # Year × Country pivot
    ws_sum.append([])
    ws_sum.append([])
    all_years = sorted(set(int(c['year']) for c in all_cases if c.get('year') and str(c['year']).isdigit()))
    ws_sum.append(['Year', 'Brazil 2016+', 'Brazil Hist.', 'Canada', 'Netherlands', 'Total'])
    header_row = ws_sum.max_row
    for col in range(1, 7):
        c = ws_sum.cell(row=header_row, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2C3E50')
        c.alignment = Alignment(horizontal='center')
    ws_sum.column_dimensions['D'].width = 14
    ws_sum.column_dimensions['E'].width = 14
    ws_sum.column_dimensions['F'].width = 10

    # Track Brazil modern vs historical separately by source
    year_br_mod  = Counter()
    year_br_hist = Counter()
    year_ca      = Counter()
    year_nl      = Counter()
    for case in brazil_cases:
        yr = case.get('year', '')
        if yr and str(yr).isdigit(): year_br_mod[int(yr)] += 1
    for case in tjsp_hist_cases:
        yr = case.get('year', '')
        if yr and str(yr).isdigit(): year_br_hist[int(yr)] += 1
    for case in canada_cases:
        yr = case.get('year', '')
        if yr and str(yr).isdigit(): year_ca[int(yr)] += 1
    for case in nl_cases:
        yr = case.get('year', '')
        if yr and str(yr).isdigit(): year_nl[int(yr)] += 1

    for yr in all_years:
        br  = year_br_mod.get(yr, 0)
        brh = year_br_hist.get(yr, 0)
        ca  = year_ca.get(yr, 0)
        nl  = year_nl.get(yr, 0)
        ws_sum.append([yr, br, brh, ca, nl, br + brh + ca + nl])

    # Move Summary to front
    wb.move_sheet('Summary', offset=-(len(wb.sheetnames)-1))

    wb.save(OUT_XLSX)
    print(f'XLSX saved: {OUT_XLSX}')

except Exception as e:
    import traceback
    print(f'XLSX error: {e}')
    traceback.print_exc()

# ── Console summary ───────────────────────────────────────────────────────────
print('\n' + '='*55)
print(f'{"GLOBAL WATER LAW DATASET":^55}')
print('='*55)
all_groups = {
    'Brazil (2016-2026)':             brazil_cases,
    'Brazil Historical TJSP (1997-2015)': tjsp_hist_cases,
    'Canada':                         canada_cases,
    'Netherlands':                    nl_cases,
}
for label, cases in all_groups.items():
    print(f'\n{label}: {len(cases)} cases')
    yr_count = Counter(str(c['year']) for c in cases if c.get('year'))
    for yr in sorted(yr_count):
        bar = '█' * min(yr_count[yr] // 10, 40)
        print(f'  {yr}: {yr_count[yr]:4d} {bar}')
print(f'\n{"GRAND TOTAL":>20}: {total} cases')
